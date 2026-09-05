import streamlit as st
import openpyxl
import json
import re
from google import genai
from google.genai import types

st.set_page_config(page_title="Financial Model Checker", layout="wide", page_icon="📊")

st.title("📊 Financial Model Sanity-Checker")
st.markdown("AI-powered audit in seconds. Upload your model, get plain-language findings.")

# --- 1. UI: Upload & Config ---
api_key = st.text_input("Gemini API Key (optional)", type="password", help="Needed for AI insights")
uploaded_file = st.file_uploader("Upload Excel Model", type=['xlsx'])

# --- 2. Deterministic Analysis Engine ---
def analyze_structure(file_bytes):
    # Load workbook using openpyxl. data_only=False ensures we read the formula strings
    wb = openpyxl.load_workbook(file_bytes, data_only=False)
    findings = []

    total_cells = 0
    formula_cells = 0
    error_values = ['#REF!', '#DIV/0!', '#VALUE!', '#N/A', '#NAME?', '#NUM!']

    # 1st Pass: Simple checks
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                total_cells += 1

                val_str = str(cell.value).upper()

                # Check for formula
                if val_str.startswith('='):
                    formula_cells += 1

                    # Broken External Links
                    has_external_link = '[' in val_str and ']' in val_str
                    if has_external_link:
                        findings.append({
                            "severity": "info",
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "title": "External Link",
                            "description": "Formula references an external workbook."
                        })

                    # Missing Sheet references (case-insensitive, strips the trailing '!';
                    # skipped for external-link formulas so their bracket syntax isn't
                    # mis-parsed as a broken in-workbook sheet reference)
                    sheet_names_upper = [s.upper() for s in wb.sheetnames]
                    sheet_refs = re.findall(r"'?([A-Z0-9_ ]+)'?!", val_str) if not has_external_link else []
                    for ref in sheet_refs:
                        if ref.upper() not in sheet_names_upper:
                            findings.append({
                                "severity": "critical",
                                "sheet": sheet_name,
                                "cell": cell.coordinate,
                                "title": "Broken Sheet Reference",
                                "description": f"References missing sheet: {ref}"
                            })
                else:
                    # Check for error strings
                    if val_str in error_values:
                        findings.append({
                            "severity": "critical",
                            "sheet": sheet_name,
                            "cell": cell.coordinate,
                            "title": "Error Value",
                            "description": f"Cell evaluates to {val_str}"
                        })

    summary = {
        "sheets": wb.sheetnames,
        "total_cells": total_cells,
        "formula_cells": formula_cells,
        "critical": len([f for f in findings if f['severity'] == 'critical']),
        "warnings": len([f for f in findings if f['severity'] == 'warning']),
        "info": len([f for f in findings if f['severity'] == 'info']),
    }

    return findings, summary

# --- 3. Gemini AI Engine ---
def run_ai_analysis(findings, summary, key):
    client = genai.Client(api_key=key)

    prompt = f"""
    You are a senior financial auditor. Review these structural findings from an Excel model.
    Workbook summary: {json.dumps(summary)}
    Findings: {json.dumps(findings)}

    Return a JSON object:
    {{
        "healthScore": 0-100,
        "executiveSummary": "Short text",
        "recommendations": ["rec 1", "rec 2"],
        "patterns": ["pattern 1"],
        "findings": [
           {{ "cell": "A1", "aiExplanation": "Why this matters", "impact": "Impact", "suggestedFix": "Fix" }}
        ]
    }}
    """

    # Use gemini-3.6-flash and force JSON output schema
    response = client.models.generate_content(
        model='gemini-3.6-flash',
        contents=prompt,
        config=types.GenerateContentConfig(
            temperature=0.2,
            response_mime_type="application/json"
        )
    )

    return json.loads(response.text)

# --- 4. Main Execution Flow ---
if uploaded_file and st.button("Analyze Model"):
    with st.spinner("Parsing workbook & running structural checks..."):
        findings, summary = analyze_structure(uploaded_file)

    ai_report = None
    if api_key:
        with st.spinner("Running AI analysis..."):
            try:
                ai_report = run_ai_analysis(findings, summary, api_key)
            except Exception as e:
                st.error(f"AI Analysis Failed: {e}")

    # --- 5. UI: Rendering the Report ---
    st.success("Analysis Complete!")

    # Scorecard
    col1, col2, col3, col4 = st.columns(4)
    health_score = ai_report['healthScore'] if ai_report else max(0, 100 - summary['critical']*15 - summary['warnings']*5)

    col1.metric("Health Score", f"{health_score}/100")
    col2.metric("Critical Issues", summary['critical'])
    col3.metric("Warnings", summary['warnings'])
    col4.metric("Info Items", summary['info'])

    # AI Summaries
    if ai_report:
        st.info(ai_report['executiveSummary'], icon="🧠")

        col_rec, col_pat = st.columns(2)
        with col_rec:
            st.subheader("Recommendations")
            for r in ai_report.get('recommendations', []):
                st.markdown(f"- {r}")
        with col_pat:
            st.subheader("Systemic Patterns")
            for p in ai_report.get('patterns', []):
                st.markdown(f"- {p}")

    st.divider()

    # Findings Cards (using st.expander)
    st.subheader("Detailed Findings")

    # Merge AI insights into structural findings
    ai_lookup = {f['cell']: f for f in ai_report['findings']} if ai_report else {}

    if not findings:
        st.write("No structural issues found!")

    for f in findings:
        ai_data = ai_lookup.get(f['cell'], {})
        icon = "🔴" if f['severity'] == 'critical' else "🟠" if f['severity'] == 'warning' else "🔵"

        with st.expander(f"{icon} **{f['sheet']}!{f['cell']}** - {f['title']}"):
            st.write(f['description'])

            if ai_data:
                st.markdown(f"**AI Explanation:** {ai_data.get('aiExplanation')}")
                st.markdown(f"**Impact:** {ai_data.get('impact')}")
                st.markdown(f"**Suggested Fix:** `{ai_data.get('suggestedFix')}`")
